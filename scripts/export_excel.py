#!/usr/bin/env python3
"""rankings.json からExcelファイルを生成するスクリプト（モバイル閲覧用）

Usage:
    python -m scripts.export_excel
"""

import json
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

PROJECT_ROOT = Path(__file__).parent.parent
WEB_DATA_DIR = PROJECT_ROOT / "web" / "data"
OUTPUT_DIR = PROJECT_ROOT / "output"


def _style_and_freeze(ws):
    """ヘッダー行を太字にし、列幅を自動調整し、ヘッダー行を固定する"""
    # 太字ヘッダー
    bold_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = bold_font
        cell.alignment = Alignment(horizontal="center")

    # 列幅自動調整
    for col_idx, col_cells in enumerate(ws.columns, 1):
        max_len = 0
        for cell in col_cells:
            val = str(cell.value or "")
            char_len = sum(2 if ord(c) > 127 else 1 for c in val)
            max_len = max(max_len, char_len)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 60)

    # ヘッダー行固定
    ws.freeze_panes = "A2"


def main():
    rankings_path = WEB_DATA_DIR / "rankings.json"
    if not rankings_path.exists():
        print(f"エラー: {rankings_path} が見つかりません。先に generate_stats.py を実行してください。")
        sys.exit(1)

    with open(rankings_path, encoding="utf-8") as f:
        data = json.load(f)

    wb = Workbook()

    # 1. 全体統計
    ws = wb.active
    ws.title = "全体統計"
    overall = data["overall"]
    ws.append(["項目", "値"])
    ws.append(["総動画数", overall["total_videos"]])
    ws.append(["メンバー数", overall["total_members"]])
    ws.append(["総演奏曲数", overall["total_songs"]])
    ws.append(["ユニーク曲数", overall["unique_songs"]])
    ws.append(["アーティスト数", overall["total_artists"]])
    _style_and_freeze(ws)

    # 2. バンド数ランキング
    ws = wb.create_sheet("バンド数ランキング")
    ws.append(["順位", "名前", "バンド数"])
    for i, item in enumerate(data["by_band_count"], 1):
        ws.append([i, item["name"], item["count"]])
    _style_and_freeze(ws)

    # 3. 曲数ランキング
    ws = wb.create_sheet("曲数ランキング")
    ws.append(["順位", "名前", "曲数"])
    for i, item in enumerate(data["by_song_count"], 1):
        ws.append([i, item["name"], item["count"]])
    _style_and_freeze(ws)

    # 4. 多様性ランキング
    ws = wb.create_sheet("多様性ランキング")
    ws.append(["順位", "名前", "ユニークアーティスト数"])
    for i, item in enumerate(data["by_artist_diversity"], 1):
        ws.append([i, item["name"], item["unique_artists"]])
    _style_and_freeze(ws)

    # 5. 人気曲
    ws = wb.create_sheet("人気曲")
    ws.append(["順位", "曲名", "アーティスト", "演奏回数"])
    for i, item in enumerate(data["popular_songs"], 1):
        ws.append([i, item["title"], item["artist"], item["play_count"]])
    _style_and_freeze(ws)

    # 6. 人気アーティスト
    ws = wb.create_sheet("人気アーティスト")
    ws.append(["順位", "アーティスト", "曲数", "バンド数"])
    for i, item in enumerate(data["popular_artists"], 1):
        ws.append([i, item["artist"], item["song_count"], item["band_count"]])
    _style_and_freeze(ws)

    # 7. ベストコンビ
    ws = wb.create_sheet("ベストコンビ")
    ws.append(["順位", "ペア", "共演回数"])
    for i, item in enumerate(data["frequent_pairs"], 1):
        pair_str = f"{item['pair'][0]} & {item['pair'][1]}"
        ws.append([i, pair_str, item["count"]])
    _style_and_freeze(ws)

    # 8. 視聴回数
    ws = wb.create_sheet("視聴回数")
    ws.append(["順位", "バンド名/動画", "視聴回数", "URL"])
    for i, item in enumerate(data["most_viewed"], 1):
        label = item.get("band_name") or item.get("title", "")
        ws.append([i, label, item["view_count"], item.get("url", "")])
    _style_and_freeze(ws)

    # 9. 視聴回数メンバー
    ws = wb.create_sheet("視聴回数メンバー")
    ws.append(["順位", "名前", "総視聴回数", "動画数"])
    for i, item in enumerate(data["view_count_members"], 1):
        ws.append([i, item["name"], item["total_views"], item["video_count"]])
    _style_and_freeze(ws)

    # 10. トリ率
    ws = wb.create_sheet("トリ率")
    ws.append(["順位", "名前", "トリ回数", "参加イベント数", "トリ率(%)"])
    for i, item in enumerate(data["tori_ranking"], 1):
        tori_pct = round(item["tori_rate"] * 100, 1)
        ws.append([i, item["name"], item["tori_count"], item["event_count"], tori_pct])
    _style_and_freeze(ws)

    # 11. イベント一覧
    ws = wb.create_sheet("イベント一覧")
    ws.append(["日付", "イベント", "バンド数", "曲数", "参加者", "アーティスト", "総視聴"])
    for item in data["event_stats"]:
        ws.append([
            item.get("date", ""),
            item["event"],
            item["bands"],
            item["songs"],
            item["members"],
            item["artists"],
            item["total_views"],
        ])
    _style_and_freeze(ws)

    # 12. パート別
    ws = wb.create_sheet("パート別")
    ws.append(["パート", "延べ出演数", "人数"])
    for item in data["part_stats"]:
        ws.append([item["part"], item["total_appearances"], item["unique_members"]])
    _style_and_freeze(ws)

    # 保存
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    output_path = OUTPUT_DIR / "keion_stats.xlsx"
    wb.save(output_path)
    print(f"Excel出力完了: {output_path}")
    print(f"  シート数: {len(wb.sheetnames)}")
    for name in wb.sheetnames:
        print(f"    - {name}")


if __name__ == "__main__":
    main()
