"""Excel出力"""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .models import ParsedVideo, Rankings
from .stats import StatsCalculator

HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def _style_header(ws):
    """ヘッダー行のスタイル設定"""
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")


def _auto_width(ws):
    """列幅を自動調整"""
    for col_idx, col in enumerate(ws.columns, 1):
        max_len = 0
        for cell in col:
            val = str(cell.value or "")
            # 日本語は2文字分として計算
            char_len = sum(2 if ord(c) > 127 else 1 for c in val)
            max_len = max(max_len, char_len)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 60)


def _add_autofilter(ws):
    """オートフィルターを設定"""
    if ws.max_row and ws.max_row > 1:
        ws.auto_filter.ref = ws.dimensions


def export_excel(
    videos: list[ParsedVideo],
    calculator: StatsCalculator,
    rankings: Rankings,
    output_path: Path,
):
    """統計データをExcelファイルに出力"""
    wb = Workbook()

    # --- シート1: 全動画一覧 ---
    ws = wb.active
    ws.title = "全動画一覧"
    ws.append(["日付", "動画タイトル", "バンド名", "曲数", "メンバー数", "曲名", "アーティスト", "メンバー(パート)", "URL"])
    for v in sorted(videos, key=lambda x: (x.date or ""), reverse=True):
        songs_str = " / ".join(f"{s.title}({s.artist})" for s in v.songs)
        artists_str = ", ".join(sorted({s.artist for s in v.songs}))
        members_str = ", ".join(
            f"{m.name}({m.part})" if m.part else m.name for m in v.members
        )
        ws.append([
            v.date.isoformat() if v.date else "",
            v.title,
            v.band_name,
            len(v.songs),
            len(v.members),
            songs_str,
            artists_str,
            members_str,
            v.url,
        ])
    _style_header(ws)
    _auto_width(ws)
    _add_autofilter(ws)

    # --- シート2: メンバー一覧 ---
    ws = wb.create_sheet("メンバー一覧")
    ws.append(["名前", "バンド数", "曲数", "ユニークアーティスト数", "学年"])
    for name in calculator.get_all_member_names():
        s = calculator.member_summary(name)
        ws.append([name, s.total_bands, s.total_songs, s.unique_artists, ", ".join(s.grades_seen)])
    _style_header(ws)
    _auto_width(ws)
    _add_autofilter(ws)

    # --- シート3: 曲一覧 ---
    ws = wb.create_sheet("曲一覧")
    ws.append(["曲名", "アーティスト", "演奏回数"])
    for item in rankings.popular_songs:
        ws.append([item["title"], item["artist"], item["play_count"]])
    _style_header(ws)
    _auto_width(ws)
    _add_autofilter(ws)

    # --- シート4: アーティスト一覧 ---
    ws = wb.create_sheet("アーティスト一覧")
    ws.append(["アーティスト", "曲数", "バンド数"])
    for item in rankings.popular_artists:
        ws.append([item["artist"], item["song_count"], item["band_count"]])
    _style_header(ws)
    _auto_width(ws)
    _add_autofilter(ws)

    # --- シート5: バンド数ランキング ---
    ws = wb.create_sheet("バンド数ランキング")
    ws.append(["順位", "名前", "バンド数"])
    for i, item in enumerate(rankings.by_band_count, 1):
        ws.append([i, item["name"], item["count"]])
    _style_header(ws)
    _auto_width(ws)

    # --- シート6: 曲数ランキング ---
    ws = wb.create_sheet("曲数ランキング")
    ws.append(["順位", "名前", "曲数"])
    for i, item in enumerate(rankings.by_song_count, 1):
        ws.append([i, item["name"], item["count"]])
    _style_header(ws)
    _auto_width(ws)

    # --- シート7: 多様性ランキング ---
    ws = wb.create_sheet("多様性ランキング")
    ws.append(["順位", "名前", "ユニークアーティスト数"])
    for i, item in enumerate(rankings.by_artist_diversity, 1):
        ws.append([i, item["name"], item["unique_artists"]])
    _style_header(ws)
    _auto_width(ws)

    # --- シート8: ベストコンビ ---
    ws = wb.create_sheet("ベストコンビ")
    ws.append(["順位", "ペア1", "ペア2", "共演回数"])
    for i, item in enumerate(rankings.frequent_pairs, 1):
        ws.append([i, item["pair"][0], item["pair"][1], item["count"]])
    _style_header(ws)
    _auto_width(ws)

    # --- シート9: 交流度ランキング ---
    ws = wb.create_sheet("交流度ランキング")
    ws.append(["順位", "名前", "共演者数", "出演バンド数"])
    for i, item in enumerate(rankings.collaboration_diversity, 1):
        ws.append([i, item["name"], item["unique_partners"], item["video_count"]])
    _style_header(ws)
    _auto_width(ws)

    # 保存
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
